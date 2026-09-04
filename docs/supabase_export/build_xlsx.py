"""Erzeugt aus dem Supabase-Export (ligas_raw.json) eine Excel-Datei mit
zwei Sheets:

  * "Spiele"  – alle Einzel-/Doppelspiele, nach jeder Begegnung eine Leerzeile
  * "Tabelle" – aktuelle Tabelle, exakt nach der Dart-Logik berechnet

Punkt-/Tabellenregeln (aus lib/model/*.dart):
  Satz: 6:x => Sieg (2:0 Satzpunkte), 5:5 => Unentschieden (1:1)
  Begegnung: mehr Satzpunkte => 2 Ligapunkte, gleich => je 1, sonst 0
             (nur wenn ALLE Spiele abgeschlossen sind)
  Tabelle:   sortiert nach Ligapunkte, Punktedifferenz, Tordifferenz,
             direkter Vergleich
"""

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SLOT_LABELS = ["D1", "E1", "D2", "E2", "D3", "E3", "D4"]


def lade_liga():
    rows = json.load(open("ligas_raw.json", encoding="utf-8"))
    return rows[0]["daten"]


def baue_indizes(liga):
    team_name = {}
    spieler_name = {}
    spieler_team = {}
    for team in liga["teams"]:
        team_name[team["id"]] = team["name"]
        for sp in team.get("spieler", []):
            voll = f"{sp.get('vorname', '').strip()} {sp.get('nachname', '').strip()}"
            spieler_name[sp["id"]] = voll.strip()
            spieler_team[sp["id"]] = team["id"]
    return team_name, spieler_name, spieler_team


# ── Satz-/Spiel-Logik (1:1 aus Dart) ─────────────────────────────────


def satz_abgeschlossen(s):
    h, g = s["heim"], s["gast"]
    return (h == 6 and g < 6) or (g == 6 and h < 6) or (h == 5 and g == 5)


def satz_punkte(s):
    h, g = s["heim"], s["gast"]
    if h == 5 and g == 5:
        return 1, 1
    if h == 6 and g < 6:
        return 2, 0
    if g == 6 and h < 6:
        return 0, 2
    return 0, 0


def spiel_saetze(spiel):
    if spiel["type"] == "einzel":
        return [spiel["satz"]] if spiel.get("satz") else []
    return spiel.get("saetze", [])


def spiel_abgeschlossen(spiel):
    saetze = spiel_saetze(spiel)
    if spiel["type"] == "einzel":
        return len(saetze) == 1 and satz_abgeschlossen(saetze[0])
    return len(saetze) == 2 and all(satz_abgeschlossen(s) for s in saetze)


def spiel_punkte(spiel):
    ph = pg = 0
    for s in spiel_saetze(spiel):
        a, b = satz_punkte(s)
        ph += a
        pg += b
    return ph, pg


def spiel_tore(spiel):
    th = tg = 0
    for s in spiel_saetze(spiel):
        th += s["heim"]
        tg += s["gast"]
    return th, tg


# ── Begegnung-Logik ──────────────────────────────────────────────────


def begegnung_spiele(beg):
    return [sp for sp in beg.get("spiele", []) if sp]


def begegnung_abgeschlossen(beg):
    spiele = beg.get("spiele", [])
    if not spiele:
        return False
    return all(sp is not None and spiel_abgeschlossen(sp) for sp in spiele)


def begegnung_satzpunkte(beg):
    ph = pg = 0
    for sp in begegnung_spiele(beg):
        a, b = spiel_punkte(sp)
        ph += a
        pg += b
    return ph, pg


def begegnung_tore(beg):
    th = tg = 0
    for sp in begegnung_spiele(beg):
        a, b = spiel_tore(sp)
        th += a
        tg += b
    return th, tg


def begegnung_ligapunkte(beg):
    if not begegnung_abgeschlossen(beg):
        return 0, 0
    ph, pg = begegnung_satzpunkte(beg)
    if ph > pg:
        return 2, 0
    if ph == pg:
        return 1, 1
    return 0, 2


# ── Tabelle ──────────────────────────────────────────────────────────


def tabelle(liga, begegnungen):
    stats = {
        t["id"]: dict(
            name=t["name"],
            lp=0,
            sp=0,
            s=0,
            u=0,
            n=0,
            tore=0,
            gegen=0,
            pf=0,
            pa=0,
            spiele=0,
        )
        for t in liga["teams"]
    }

    for beg in begegnungen:
        if not begegnung_abgeschlossen(beg):
            continue
        heim, gast = beg["heim"], beg["gast"]
        lph, lpg = begegnung_ligapunkte(beg)
        th, tg = begegnung_tore(beg)
        sph, spg = begegnung_satzpunkte(beg)

        for tid, lp, tf, ta, pf, pa in (
            (heim, lph, th, tg, sph, spg),
            (gast, lpg, tg, th, spg, sph),
        ):
            st = stats[tid]
            st["lp"] += lp
            st["spiele"] += 1
            st["tore"] += tf
            st["gegen"] += ta
            st["pf"] += pf
            st["pa"] += pa

        # Satz-Bilanz (S/U/N) je Team, satzweise wie im Dart-Code
        for sp in begegnung_spiele(beg):
            for s in spiel_saetze(sp):
                if not satz_abgeschlossen(s):
                    continue
                ph, pg = satz_punkte(s)
                # Heim
                if ph == 2:
                    stats[heim]["s"] += 1
                elif ph == 0:
                    stats[heim]["n"] += 1
                else:
                    stats[heim]["u"] += 1
                # Gast
                if pg == 2:
                    stats[gast]["s"] += 1
                elif pg == 0:
                    stats[gast]["n"] += 1
                else:
                    stats[gast]["u"] += 1

    # direkter Vergleich
    def direkter_vergleich(a_id, b_id):
        pa = pb = 0
        for beg in begegnungen:
            if not begegnung_abgeschlossen(beg):
                continue
            lph, lpg = begegnung_ligapunkte(beg)
            if beg["heim"] == a_id and beg["gast"] == b_id:
                pa += lph
                pb += lpg
            elif beg["heim"] == b_id and beg["gast"] == a_id:
                pb += lph
                pa += lpg
        return pa - pb

    import functools

    def cmp(x, y):
        if y["lp"] != x["lp"]:
            return y["lp"] - x["lp"]
        if (y["pf"] - y["pa"]) != (x["pf"] - x["pa"]):
            return (y["pf"] - y["pa"]) - (x["pf"] - x["pa"])
        if (y["tore"] - y["gegen"]) != (x["tore"] - x["gegen"]):
            return (y["tore"] - y["gegen"]) - (x["tore"] - x["gegen"])
        # direkter Vergleich: b vs a (s. Dart _direkterVergleich(b,a))
        return direkter_vergleich(y["id"], x["id"])

    for tid, st in stats.items():
        st["id"] = tid

    return sorted(stats.values(), key=functools.cmp_to_key(cmp))


# ── Styling-Helfer ───────────────────────────────────────────────────

FONT = "Calibri"  # serifenlos
GROESSE = 14
GROESSE_KOPF = 15

DUNKEL = PatternFill("solid", fgColor="2E1A47")
KOPF = PatternFill("solid", fgColor="5E35B1")
BEG = PatternFill("solid", fgColor="D1C4E9")
ZEBRA = PatternFill("solid", fgColor="F3EEFB")
WEISS = Font(name=FONT, size=GROESSE_KOPF, color="FFFFFF", bold=True)
FETT = Font(name=FONT, size=GROESSE, bold=True)


def setze_font(ws):
    """Setzt auf allen Zellen den serifenlosen Font in der Zielgröße und
    behält dabei Fettung/Farbe der bereits gesetzten Fonts bei."""
    for row in ws.iter_rows():
        for z in row:
            alt = z.font
            groesse = alt.size if (alt and alt.size and alt.size > GROESSE) else GROESSE
            z.font = Font(
                name=FONT,
                size=groesse,
                bold=bool(alt and alt.bold),
                color=alt.color if alt else None,
            )
ZENTRIERT = Alignment(horizontal="center", vertical="center")
LINKS = Alignment(horizontal="left", vertical="center")
DUENN = Side(style="thin", color="C9BCE0")
RAHMEN = Border(left=DUENN, right=DUENN, top=DUENN, bottom=DUENN)


def saetze_text(spiel):
    return "  ".join(f"{s['heim']}:{s['gast']}" for s in spiel_saetze(spiel))


# ── Sheet: Spiele ────────────────────────────────────────────────────


def sheet_spiele(wb, liga, begegnungen, team_name, spieler_name):
    ws = wb.active
    ws.title = "Spiele"
    kopf = [
        "Slot",
        "Heim-Aufstellung",
        "Tore",
        "Gast-Aufstellung",
        "Satzpkt.",
    ]
    ncols = len(kopf)
    ws.append(kopf)
    for c in range(1, ncols + 1):
        z = ws.cell(row=1, column=c)
        z.fill = KOPF
        z.font = WEISS
        z.alignment = ZENTRIERT

    def name(pid):
        return spieler_name.get(pid, pid or "–")

    for beg in begegnungen:
        st_nr = beg["_st_nummer"]
        heim_n = team_name.get(beg["heim"], beg["heim"])
        gast_n = team_name.get(beg["gast"], beg["gast"])
        sph, spg = begegnung_satzpunkte(beg)
        th, tg = begegnung_tore(beg)
        fertig = begegnung_abgeschlossen(beg)

        # Begegnungs-Kopfzeile: Spieltag im Slot, Teamnamen über den
        # Spielern, Tore-Summe in der Tore-Spalte, Satzpunkt-Summe ganz rechts
        st_label = f"ST {st_nr}" if fertig else f"ST {st_nr}*"
        ws.append([st_label, heim_n, f"{th}:{tg}", gast_n, f"{sph}:{spg}"])
        r = ws.max_row
        for c in range(1, ncols + 1):
            z = ws.cell(row=r, column=c)
            z.fill = BEG
            z.font = FETT
            z.alignment = LINKS if c in (2, 4) else ZENTRIERT

        # Pro Spiel: jeder Satz eine eigene Zeile.
        # Bei Doppeln zeigt nur die erste Satz-Zeile Slot + Namen.
        for idx, spiel in enumerate(beg.get("spiele", [])):
            slot = SLOT_LABELS[idx] if idx < len(SLOT_LABELS) else f"#{idx + 1}"
            if not spiel:
                ws.append([slot, "—", "", "—", ""])
                continue
            if spiel["type"] == "doppel":
                heim_a = " / ".join(name(p) for p in spiel.get("heim", []))
                gast_a = " / ".join(name(p) for p in spiel.get("gast", []))
            else:
                heim_a = name(spiel.get("heim"))
                gast_a = name(spiel.get("gast"))

            saetze = spiel_saetze(spiel)
            if not saetze:
                ws.append([slot, heim_a, "", gast_a, ""])
                rr = ws.max_row
                for c in (1, 3, 5):
                    ws.cell(row=rr, column=c).alignment = ZENTRIERT
                continue
            for satz_idx, s in enumerate(saetze):
                ph, pg = satz_punkte(s)
                erste = satz_idx == 0
                ws.append(
                    [
                        slot if erste else "",
                        heim_a if erste else "",
                        f"{s['heim']}:{s['gast']}",
                        gast_a if erste else "",
                        f"{ph}:{pg}",
                    ]
                )
                rr = ws.max_row
                for c in (1, 3, 5):
                    ws.cell(row=rr, column=c).alignment = ZENTRIERT

        ws.append([])  # Leerzeile nach jeder Begegnung

    breiten = [8, 34, 10, 34, 10]
    for i, b in enumerate(breiten, start=1):
        ws.column_dimensions[get_column_letter(i)].width = b
    ws.freeze_panes = "A2"
    setze_font(ws)


# ── Sheet: Tabelle ───────────────────────────────────────────────────


def sheet_tabelle(wb, tab):
    ws = wb.create_sheet("Tabelle")
    kopf = [
        "Pl.",
        "Team",
        "Sp",
        "S",
        "U",
        "N",
        "Tore",
        "Diff",
        "Satzpkt",
        "S-Diff",
        "Punkte",
    ]
    ws.append(kopf)
    for c in range(1, len(kopf) + 1):
        z = ws.cell(row=1, column=c)
        z.fill = KOPF
        z.font = WEISS
        z.alignment = ZENTRIERT
        z.border = RAHMEN

    for i, t in enumerate(tab, start=1):
        ws.append(
            [
                i,
                t["name"],
                t["spiele"],
                t["s"],
                t["u"],
                t["n"],
                f"{t['tore']}:{t['gegen']}",
                t["tore"] - t["gegen"],
                f"{t['pf']}:{t['pa']}",
                t["pf"] - t["pa"],
                t["lp"],
            ]
        )
        r = ws.max_row
        for c in range(1, len(kopf) + 1):
            z = ws.cell(row=r, column=c)
            z.border = RAHMEN
            z.alignment = LINKS if c == 2 else ZENTRIERT
            if i % 2 == 0:
                z.fill = ZEBRA
        ws.cell(row=r, column=11).font = FETT

    breiten = [5, 26, 6, 6, 6, 6, 10, 8, 10, 9, 9]
    for i, b in enumerate(breiten, start=1):
        ws.column_dimensions[get_column_letter(i)].width = b
    ws.freeze_panes = "A2"
    setze_font(ws)


# ── Main ─────────────────────────────────────────────────────────────


def main():
    liga = lade_liga()
    team_name, spieler_name, _ = baue_indizes(liga)

    # Begegnungen mit Spieltag-Nummer anreichern, in Spielplan-Reihenfolge
    begegnungen = []
    for tag in liga.get("hinrunde", []) + liga.get("rueckrunde", []):
        for beg in tag.get("begegnungen", []):
            beg = dict(beg)
            beg["_st_nummer"] = tag.get("nummer")
            begegnungen.append(beg)

    wb = Workbook()
    sheet_spiele(wb, liga, begegnungen, team_name, spieler_name)
    sheet_tabelle(wb, tabelle(liga, begegnungen))

    datei = "ue45-liga.xlsx"
    wb.save(datei)

    # Kurz-Report
    fertig = sum(1 for b in begegnungen if begegnung_abgeschlossen(b))
    print(f"Datei: {datei}")
    print(f"Begegnungen gesamt: {len(begegnungen)}  (abgeschlossen: {fertig})")
    print("\nTabelle:")
    for i, t in enumerate(tabelle(liga, begegnungen), start=1):
        print(
            f"  {i}. {t['name']:<22} {t['lp']:>2} Pkt"
            f"  | Sp {t['spiele']}  Tore {t['tore']}:{t['gegen']}"
            f"  Satzpkt {t['pf']}:{t['pa']}"
        )


if __name__ == "__main__":
    main()
